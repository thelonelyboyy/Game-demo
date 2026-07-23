"""Write edited game_data.xlsx back into the corresponding game data files.

Run:  python scripts/xlsx_to_game_data.py
然后在 Godot 编辑器里导入一次（--editor --quit）。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import card_table as ct
import card_effect_table as cet
import fusion_table as ft
import game_data_labels as labels
import blessing_table as bt
import ai_table as at
import enemy_table as ent
import level_table as lt
import event_table as et
import potion_table as pt
import shop_price_table as st
import relic_table as rt

XLSX = Path(__file__).resolve().parents[1] / "game_data.xlsx"


def _rowvals(ws):
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        yield idx, [c.value for c in ws[r]]


def import_cards(ws, errors) -> int:
    changed = 0
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        path = ct.PROJECT_ROOT / rel
        if not path.exists():
            errors.append(f"[卡牌] file not found {rel}")
            continue

        def gv(name, default=""):
            v = vals[idx[name]]
            return default if v is None else v

        try:
            row = {
                "name": str(gv("名称")),
                "cost": int(gv("费用", 0)),
                "type": str(gv("类型")).strip(),
                "target": str(gv("目标")).strip(),
                "rarity": str(gv("稀有度")).strip(),
                "exhausts": str(gv("消耗")).strip() == "是",
                "element": str(gv("元素")).strip(),
                "upgrade": str(gv("突破方向")).strip(),
                "growth_trigger": str(gv("成长触发", "无")).strip(),
                "growth_amount": int(gv("每次成长值", 1)),
                "growth_limit": int(gv("成长上限", 0)),
                "effect_amounts": [],
            }
            for key, column in (
                ("exhausts", "消耗"), ("retains", "保留"), ("innate", "固有"),
                ("eternal", "永恒"), ("ethereal", "虚无"), ("temporary_keyword", "临时"),
                ("cyclic", "周天"), ("unplayable", "不可打出"),
                ("status_card", "状态牌"), ("curse_card", "诅咒牌"),
            ):
                row[key] = str(gv(column, "否")).strip() == "是"
            for key, enabled_column, count_column in (
                ("search_count", "检索", "检索张数"),
                ("retrieve_count", "取回", "取回张数"),
                ("reclaim_count", "归墟", "归墟张数"),
            ):
                enabled = str(gv(enabled_column, "否")).strip() == "是"
                count = int(gv(count_column, 1 if enabled else 0))
                if enabled and not 1 <= count <= 10:
                    raise ValueError(f"{count_column}必须在 1 到 10 之间")
                row[key] = max(count, 1) if enabled else 0
            for key, table in (("type", ct.TYPE_R), ("target", ct.TARGET_R),
                               ("rarity", ct.RARITY_R), ("element", ct.ELEMENT_R),
                               ("upgrade", ct.UPGRADE_R), ("growth_trigger", ct.GROWTH_R)):
                if row[key] not in table:
                    raise ValueError(f"非法{key}值 '{row[key]}'")
        except (ValueError, TypeError) as e:
            errors.append(f"[卡牌] {rel}: {e}")
            continue

        if ct.update_card_file(path, row):
            changed += 1
    return changed


def import_relic_params(ws, errors) -> int:
    by_file: dict[str, list[dict]] = {}
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        raw_param = str(vals[idx["参数定位键"]]).strip()
        value = vals[idx["参数值"]]
        if isinstance(value, str) and value.strip() in ("是", "否"):
            value = value.strip() == "是"
        elif value is None:
            errors.append(f"[法宝参数] {rel} {raw_param}: 参数值不可留空")
            continue
        by_file.setdefault(rel, []).append({"raw_param": raw_param, "value": value})
    changed = 0
    for rel, edits in by_file.items():
        try:
            changed += int(rt.write_parameters(rel, edits))
        except (ValueError, OSError) as error:
            errors.append(f"[法宝参数] {rel}: {error}")
    return changed


def import_card_effects(ws, errors) -> int:
    by_file: dict[str, list[dict]] = {}
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        try:
            value = vals[idx["参数值"]]
            if value is None:
                raise ValueError("参数值不可留空")
            number = float(value)
            by_file.setdefault(rel, []).append({
                "effect_id": str(vals[idx["效果ID"]]).strip(),
                "param": str(vals[idx["参数名称"]]).strip(),
                "value": int(number) if number.is_integer() else number,
            })
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[卡牌效果] {rel}: {e}")
    changed = 0
    for rel, edits in by_file.items():
        try:
            changed += int(cet.write_file(rel, edits))
        except (ValueError, OSError) as e:
            errors.append(f"[卡牌效果] {rel}: {e}")
    return changed


def import_fusions(ws, errors) -> int:
    rows = []
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["配方ID"]]:
            continue
        try:
            enabled = str(vals[idx["启用"]] or "是").strip() == "是"
            rows.append({
                "enabled": enabled,
                "file": str(vals[idx["文件"]] or "").strip(),
                "recipe_id": str(vals[idx["配方ID"]]).strip(),
                "first_id": str(vals[idx["原料卡1 ID"]] or "").strip(),
                "second_id": str(vals[idx["原料卡2 ID"]] or "").strip(),
                "mode": str(vals[idx["结果方式"]] or "动态合成").strip(),
                "result_path": str(vals[idx["结果卡文件"]] or "").strip(),
                "result_name": str(vals[idx["结果名称"]] or ""),
                "result_cost": vals[idx["结果费用"]],
            })
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[融合配方] 第 {len(rows) + 2} 行: {e}")
    if errors:
        return 0
    try:
        return ft.write_rows(rows)
    except (ValueError, TypeError, KeyError, OSError) as e:
        errors.append(f"[融合配方] {e}")
        return 0


def import_blessings(ws, errors) -> bool:
    rows = []
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["来源"]]:
            continue

        def gv(name, default=""):
            v = vals[idx[name]]
            return default if v is None else v

        effects = []
        for i in range(1, bt.MAX_EFFECTS + 1):
            t = gv(f"效果{i}")
            a = vals[idx[f"效果{i}值"]]
            t = str(t).strip()
            if t:
                try:
                    effects.append((t, int(a) if a is not None else 0))
                except (ValueError, TypeError):
                    errors.append(f"[祝福] {gv('祝福名')}: 效果{i}值非数字 '{a}'")
        rows.append({
            "source": str(gv("来源")).strip(),
            "source_desc": str(gv("来源描述")),
            "name": str(gv("祝福名")),
            "desc": str(gv("祝福描述")),
            "cls": labels.CLASS_LABELS_R.get(str(gv("命格")).strip(), str(gv("命格")).strip()),
            "icon": labels.ICON_LABELS_R.get(str(gv("图标")).strip(), str(gv("图标")).strip()),
            "effects": [(labels.BLESSING_EFFECT_LABELS_R.get(t, t), a) for t, a in effects],
        })
    return bt.write_rows(rows)


def import_ai(ws, errors) -> int:
    # group by action configuration file
    by_file: dict[str, dict] = {}
    order: list[str] = []
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["行动配置文件"]]:
            continue
        rel = str(vals[idx["行动配置文件"]]).strip()
        if rel not in by_file:
            by_file[rel] = {"seq": "", "actions": []}
            order.append(rel)
        params = {}
        try:
            for i in range(1, at.MAX_PARAMS + 1):
                display_name = str(vals[idx[f"参数{i}"]] or "").strip()
                value = vals[idx[f"参数{i}值"]]
                if display_name and value is not None:
                    name = labels.PARAM_LABELS_R.get(display_name, display_name)
                    params[name] = {"name": name, "value": value}
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[怪物行动] {rel} {vals[idx['行动ID']]}: {e}")
            continue
        by_file[rel]["actions"].append({"node": str(vals[idx["行动ID"]]).strip(), "params": list(params.values())})
        seq = vals[idx["出招序列"]]
        if seq not in (None, "") and not by_file[rel]["seq"]:
            by_file[rel]["seq"] = str(seq).strip()

    changed = 0
    for rel in order:
        info = by_file[rel]
        if at.write_file(rel, info["seq"], info["actions"]):
            changed += 1
    return changed


def import_enemies(ws, errors) -> int:
    changed = 0
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        try:
            row = {"file": rel, "name": str(vals[idx["名称"]] or ""),
                   "description": str(vals[idx["介绍"]] or ""), "max_health": int(vals[idx["最大生命"]]),
                   "phase_ratio": float(vals[idx["二阶段阈值"]] or 0),
                   "phase_name": str(vals[idx["二阶段名称"]] or ""),
                   "phase_block": int(vals[idx["二阶段护体"]] or 0),
                   "phase_damage_bonus": float(vals[idx["二阶段伤害加成"]] or 0),
                   "phase_sequence": str(vals[idx["二阶段序列"]] or "")}
            if ent.write_file(row):
                changed += 1
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[怪物] {rel}: {e}")
    return changed


def import_levels(ws, errors) -> int:
    changed = 0
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        try:
            count = int(vals[idx["敌人数"]])
            enemy_ids = [str(vals[idx[f"敌人{i}"]] or "").strip() for i in range(1, count + 1)]
            row = {"file": rel, "scene": str(vals[idx["场景"]]), "tier": str(vals[idx["类型"]]),
                   "chapter_min": int(vals[idx["章节起"]]), "chapter_max": int(vals[idx["章节止"]]),
                   "weight": float(vals[idx["权重"]]), "gold_min": int(vals[idx["金币下限"]]),
                   "gold_max": int(vals[idx["金币上限"]]), "health_multiplier": float(vals[idx["生命倍率"]]),
                   "damage_multiplier": float(vals[idx["伤害倍率"]]), "enemy_ids": enemy_ids}
            if row["tier"] not in lt.TIER_R or not (1 <= row["chapter_min"] <= row["chapter_max"] <= 3):
                raise ValueError("类型或章节范围无效")
            if row["gold_min"] > row["gold_max"] or row["health_multiplier"] <= 0 or row["damage_multiplier"] <= 0:
                raise ValueError("金币区间或怪物倍率无效")
            if len(enemy_ids) != count or any(not enemy_id for enemy_id in enemy_ids):
                raise ValueError("怪物槽位不可留空")
            changed += lt.write_row(row)
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[关卡设计] {rel}: {e}")
    return changed


def import_events(ws, errors) -> int:
    changed = 0
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        choices = []
        try:
            for i in range(1, et.MAX_CHOICES + 1):
                effect = labels.event_effect_from_zh(str(vals[idx[f"选项{i}效果"]] or ""))
                text = str(vals[idx[f"选项{i}文案"]] or "")
                if (text or effect) and not et.validate_effect(effect):
                    raise ValueError(f"选项{i}效果格式无效 '{effect}'")
                choices.append({"text": text,
                                "effect": effect,
                                "amount": int(vals[idx[f"选项{i}数值"]] or 0)})
            row = {"file": rel, "title": str(vals[idx["标题"]] or ""),
                   "body": str(vals[idx["正文"]] or ""), "choices": choices}
            if et.write_event(row):
                changed += 1
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[事件] {rel}: {e}")
    return changed


def import_event_roots(ws, errors) -> int:
    changed = 0
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        try:
            texts, effects, amounts = [""], [""], [0]
            for name in et.ROOT_NAMES[1:]:
                texts.append(str(vals[idx[f"{name}文案"]] or ""))
                effects.append(labels.event_effect_from_zh(str(vals[idx[f"{name}效果"]] or "")))
                amounts.append(int(vals[idx[f"{name}数值"]] or 0))
            row = {"file": rel, "root_choice_index": int(vals[idx["覆盖选项"]]) - 1,
                   "root_texts": texts, "root_effects": effects, "root_amounts": amounts}
            if et.write_root_overrides(row):
                changed += 1
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[事件灵根] {rel}: {e}")
    return changed


def import_potions(ws, errors) -> int:
    changed = 0
    parsed = {p["file"]: p for p in pt.parse_all()}
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        source = parsed.get(rel)
        if not source:
            errors.append(f"[符箓丹药] file not found {rel}")
            continue
        try:
            row = dict(source)
            row.update({"name": str(vals[idx["名称"]] or ""), "category": str(vals[idx["类别"]] or ""),
                        "rarity": str(vals[idx["稀有度"]] or ""), "target": str(vals[idx["目标"]] or ""),
                        "out_of_combat": str(vals[idx["战斗外可用"]] or "否"),
                        "character": str(vals[idx["职业"]] or "通用"), "tooltip": str(vals[idx["说明"]] or "")})
            for i, effect in enumerate(row["effects"], 1):
                for j, param in enumerate(effect["params"], 1):
                    value = vals[idx[f"效果{i}值{j}"]]
                    if value is not None:
                        param["value"] = int(value) if float(value).is_integer() else float(value)
                    param["name"] = labels.PARAM_LABELS_R.get(
                        str(vals[idx[f"效果{i}参数{j}"]] or param["name"]).strip(), param["name"]
                    )
            if pt.write_file(row):
                changed += 1
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[符箓丹药] {rel}: {e}")
    return changed


def import_shop_prices(ws, errors) -> int:
    rows = []
    for idx, vals in _rowvals(ws):
        if not vals or not vals[idx["定位键"]]:
            continue
        try:
            rows.append({"key": str(vals[idx["定位键"]]),
                         "base": vals[idx["基价"]], "min": vals[idx["最低价"]],
                         "max": vals[idx["最高价"]], "percent": vals[idx["比例"]]})
        except (ValueError, TypeError, KeyError) as e:
            errors.append(f"[商店价格] {vals[idx['定位键']]}: {e}")
    if errors:
        return 0
    try:
        return st.write_rows(rows)
    except (ValueError, TypeError, KeyError) as e:
        errors.append(f"[商店价格] {e}")
        return 0


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"not found: {XLSX} (先跑 game_data_to_xlsx.py)")

    wb = load_workbook(XLSX, data_only=True)
    errors: list[str] = []

    n_cards = import_cards(wb["卡牌"], errors)
    n_card_effects = import_card_effects(wb["卡牌效果"], errors)
    n_fusions = import_fusions(wb["融合配方"], errors)
    bless_changed = import_blessings(wb["祝福"], errors)
    n_events = import_events(wb["事件"], errors)
    n_event_roots = import_event_roots(wb["事件灵根"], errors)
    n_potions = import_potions(wb["符箓丹药"], errors)
    n_relic_params = import_relic_params(wb["法宝参数"], errors) if "法宝参数" in wb.sheetnames else 0
    n_enemies = import_enemies(wb["怪物"], errors)
    n_ai = import_ai(wb["怪物行动"], errors)
    n_levels = import_levels(wb["关卡设计"], errors)
    n_shop_files = import_shop_prices(wb["商店价格"], errors)

    print(f"卡牌: 更新 {n_cards} 个文件")
    print(f"卡牌效果: 更新 {n_card_effects} 个文件")
    print(f"融合配方: 更新 {n_fusions} 个文件")
    print(f"祝福: {'已更新 blessings.json' if bless_changed else '无改动'}")
    print(f"事件: 更新 {n_events} 个文件")
    print(f"事件灵根: 更新 {n_event_roots} 个文件")
    print(f"符箓丹药: 更新 {n_potions} 个文件")
    print(f"法宝参数: 更新 {n_relic_params} 个文件")
    print(f"怪物: 更新 {n_enemies} 个文件")
    print(f"怪物行动: 更新 {n_ai} 个文件")
    print(f"关卡设计: 更新 {n_levels} 个资源/场景")
    print(f"商店价格: 更新 {n_shop_files} 个脚本")
    if errors:
        print("跳过(有错误):")
        for e in errors:
            print("  " + e)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
