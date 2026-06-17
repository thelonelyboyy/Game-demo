"""Export all card .tres data into cards.xlsx for balance editing.

Run:  python scripts/cards_to_xlsx.py
Then edit cards.xlsx and run scripts/xlsx_to_cards.py to write changes back.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

import card_table as ct

OUT = Path(__file__).resolve().parents[1] / "cards.xlsx"

HEADERS = ["文件", "id", "职业", "名称", "费用", "类型", "目标", "稀有度", "消耗", "元素"]
for i in range(1, ct.MAX_EFFECTS + 1):
    HEADERS += [f"效果{i}", f"效果{i}值"]

# 1-based column indices
READONLY_COLS = {1, 2, 3}  # 文件 / id / 职业
for i in range(ct.MAX_EFFECTS):
    READONLY_COLS.add(11 + i * 2)  # 效果N (type label, read-only)

FONT = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="2C2418")
READONLY_FILL = PatternFill("solid", fgColor="EDEDED")
EDIT_FILL = PatternFill("solid", fgColor="FFFDE7")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "cards"

    ws.append(HEADERS)
    for c, _ in enumerate(HEADERS, 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FONT, bold=True, color="FFF0BC")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for card in (ct.parse_card(p) for p in ct.find_card_files()):
        row = [card["file"], card["id"], card["profession"], card["name"], card["cost"],
               card["type"], card["target"], card["rarity"], "是" if card["exhausts"] else "否", card["element"]]
        for i in range(ct.MAX_EFFECTS):
            if i < len(card["effects"]):
                label, amount, extra = card["effects"][i]
                row += [f"{label}({extra})" if extra else label, amount]
            else:
                row += ["", None]
        ws.append(row)

    # styling per data cell
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c != 1 else "left", vertical="center")
            cell.fill = READONLY_FILL if c in READONLY_COLS else EDIT_FILL

    # dropdowns for the enum columns (类型/目标/稀有度/消耗/元素 -> cols 6,7,8,9,10)
    def dv(values: list[str]) -> DataValidation:
        d = DataValidation(type="list", formula1='"%s"' % ",".join(values), allow_blank=False)
        ws.add_data_validation(d)
        return d
    last = ws.max_row
    dvs = {6: list(ct.TYPE.values()), 7: list(ct.TARGET.values()), 8: list(ct.RARITY.values()),
           9: ["是", "否"], 10: list(ct.ELEMENT.values())}
    for col, vals in dvs.items():
        letter = ws.cell(1, col).column_letter
        dv(vals).add(f"{letter}2:{letter}{last}")

    widths = [42, 24, 6, 12, 5, 6, 7, 7, 5, 5] + [10, 7] * ct.MAX_EFFECTS
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "D2"

    # legend sheet
    leg = wb.create_sheet("说明")
    notes = [
        ["卡牌数据表 — 使用说明", ""],
        ["", ""],
        ["1. 只改【浅黄色】列；【灰色】列(文件/id/职业/效果类型)是只读定位用，请勿改动。", ""],
        ["2. 改完保存，运行：python scripts/xlsx_to_cards.py", ""],
        ["3. 然后在 Godot 编辑器里导入一次（或 --headless --editor --quit）。", ""],
        ["4. 效果N值 = 该效果的数值：伤害/护体/抽牌/自损/治疗的点数，状态的层数。", ""],
        ["5. 增删效果、改效果类型、改状态种类，仍需在 Godot 里改（本表只调数值）。", ""],
        ["", ""],
        ["类型", " / ".join(ct.TYPE.values())],
        ["目标", " / ".join(ct.TARGET.values())],
        ["稀有度", " / ".join(ct.RARITY.values())],
        ["元素", " / ".join(ct.ELEMENT.values())],
    ]
    for r in notes:
        leg.append(r)
    leg.column_dimensions["A"].width = 70
    leg["A1"].font = Font(name=FONT, bold=True, size=13)

    wb.save(OUT)
    print(f"wrote {OUT}  ({ws.max_row - 1} cards)")


if __name__ == "__main__":
    main()
