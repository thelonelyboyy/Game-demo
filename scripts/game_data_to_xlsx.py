"""Export tunable game data into one workbook: game_data.xlsx.

包含卡牌、卡牌效果、融合配方、祝福、事件、事件灵根、怪物、怪物行动、关卡设计、符箓丹药、商店价格与说明。
改完表后运行 scripts/xlsx_to_game_data.py 回写，再在 Godot 导入一次。
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

import card_table as ct
import card_effect_table as cet
import fusion_table as ft
import game_data_labels as labels
import blessing_table as bt
import ai_table as at
import enemy_table as ent
import level_table as lt
import event_table as et
import potion_table as pt
import shop_price_table as st

OUT = Path(__file__).resolve().parents[1] / "game_data.xlsx"

FONT = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="2C2418")
READONLY_FILL = PatternFill("solid", fgColor="EDEDED")
EDIT_FILL = PatternFill("solid", fgColor="FFFDE7")
_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FONT, bold=True, color="FFF0BC")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_body(ws, ncols, readonly_cols):
    for r in range(2, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c != 1 else "left", vertical="center")
            cell.fill = READONLY_FILL if c in readonly_cols else EDIT_FILL


def _dv(ws, values, col, last_row):
    d = DataValidation(type="list", formula1='"%s"' % ",".join(values), allow_blank=True)
    ws.add_data_validation(d)
    d.add(f"{ws.cell(1, col).column_letter}2:{ws.cell(1, col).column_letter}{last_row}")


def _dv_formula(ws, formula, col, last_row):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d)
    d.add(f"{ws.cell(1, col).column_letter}2:{ws.cell(1, col).column_letter}{last_row}")


def _widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w


def build_cards(ws):
    headers = [
        "文件", "ID", "职业", "名称", "费用", "类型", "目标", "稀有度", "元素",
        "突破方向",
        "消耗", "保留", "固有", "永恒", "虚无", "临时", "周天", "不可打出", "状态牌", "诅咒牌",
        "检索", "检索张数", "取回", "取回张数", "归墟", "归墟张数", "关键词汇总", "效果概览",
    ]
    ws.append(headers)
    readonly = {1, 2, 3, 27, 28}
    overview: dict[str, list[str]] = {}
    for effect in cet.parse_all():
        label = f'{effect["trigger"]}：{effect["effect"]}'
        if label not in overview.setdefault(effect["card_id"], []):
            overview[effect["card_id"]].append(label)
    for card in (ct.parse_card(p) for p in ct.find_card_files()):
        yn = lambda value: "是" if value else "否"
        row = [
            card["file"], card["id"], card["profession"], card["name"], card["cost"],
            card["type"], card["target"], card["rarity"], card["element"],
            card["upgrade"],
            yn(card["exhausts"]), yn(card["retains"]), yn(card["innate"]), yn(card["eternal"]),
            yn(card["ethereal"]), yn(card["temporary_keyword"]), yn(card["cyclic"]),
            yn(card["unplayable"]), yn(card["status_card"]), yn(card["curse_card"]),
            yn(card["search_count"] > 0), card["search_count"],
            yn(card["retrieve_count"] > 0), card["retrieve_count"],
            yn(card["reclaim_count"] > 0), card["reclaim_count"], card["keyword_summary"],
            "；".join(overview.get(card["id"], [])),
        ]
        ws.append(row)
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), readonly)
    last = ws.max_row
    _dv(ws, list(ct.TYPE.values()), 6, last)
    _dv(ws, list(ct.TARGET.values()), 7, last)
    _dv(ws, list(ct.RARITY.values()), 8, last)
    _dv(ws, list(ct.ELEMENT.values()), 9, last)
    _dv(ws, list(ct.UPGRADE.values()), 10, last)
    for col in (11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 25):
        _dv(ws, ["是", "否"], col, last)
    _widths(ws, [42, 24, 6, 12, 5, 6, 7, 7, 5, 16] + [7] * 16 + [30, 42])
    ws.freeze_panes = "D2"


def build_card_effects(ws):
    headers = ["文件", "卡牌ID", "卡牌名称", "触发时机", "效果ID", "效果名称", "效果脚本文件", "参数名称", "参数值"]
    ws.append(headers)
    for effect in cet.parse_all():
        ws.append([effect["file"], effect["card_id"], effect["card_name"], effect["trigger"], effect["effect_id"],
                   effect["effect"], effect["script_path"], effect["param"], effect["value"]])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), set(range(1, 9)))
    _widths(ws, [42, 22, 15, 12, 28, 22, 52, 21, 11])
    ws.freeze_panes = "D2"


def build_keyword_standards(ws):
    headers = ["关键词", "启用列", "数值列", "标准效果", "冲突与结算优先级"]
    ws.append(headers)
    rows = [
        ["消耗", "消耗", "", "打出后进入消耗堆；未打出时正常弃置。", "打出：临时 > 消耗 > 功法移出 > 周天 > 弃牌"],
        ["保留", "保留", "", "回合结束不会进入弃牌堆；使用后按其他关键词结算。", "回合末：临时 > 虚无 > 保留 > 弃牌"],
        ["固有", "固有", "", "战斗第一回合优先进入起手。", "全部固有牌先移动到抽牌堆顶"],
        ["检索X", "检索", "检索张数", "打开抽牌堆选择 X 张牌。", "检索→取回→归墟"],
        ["取回X", "取回", "取回张数", "打开弃牌堆选择 X 张牌。", "检索→取回→归墟"],
        ["归墟X", "归墟", "归墟张数", "打开消耗堆选择 X 张牌。", "检索→取回→归墟"],
        ["永恒", "永恒", "", "无法打出，不能删除、变化或合炼。", "永恒限制始终优先"],
        ["虚无", "虚无", "", "回合结束仍在手牌时进入消耗堆。", "回合末：临时 > 虚无 > 保留 > 弃牌"],
        ["临时", "临时", "", "打出后或回合结束仍在手牌时直接移除。", "打出与回合末均最高"],
        ["周天", "周天", "", "打出后置于抽牌堆顶。", "打出：临时 > 消耗 > 功法移出 > 周天 > 弃牌"],
        ["不可打出", "不可打出", "", "不能主动使用。", "与状态牌、诅咒牌、永恒共同阻止打出"],
        ["状态牌", "状态牌", "", "用于污染抽牌循环，默认不可打出。", "“可打出”机制标签可覆盖"],
        ["诅咒牌", "诅咒牌", "", "负面牌组污染，默认不可打出。", "“可打出”机制标签可覆盖"],
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), set(range(1, len(headers) + 1)))
    _widths(ws, [15, 15, 15, 52, 54])
    ws.freeze_panes = "B2"


def build_effect_standards(ws):
    headers = ["标准效果组件", "效果脚本文件", "可调数值参数", "使用卡牌数", "参数行数"]
    ws.append(headers)
    catalog: dict[tuple[str, str], dict] = defaultdict(lambda: {"params": set(), "cards": set(), "rows": 0})
    for effect in cet.parse_all():
        key = (effect["effect"], effect["script_path"])
        catalog[key]["params"].add(effect["param"])
        catalog[key]["cards"].add(effect["card_id"])
        catalog[key]["rows"] += 1
    for (effect_name, script_path), item in sorted(catalog.items()):
        ws.append([effect_name, script_path, "、".join(sorted(item["params"])), len(item["cards"]), item["rows"]])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), set(range(1, len(headers) + 1)))
    _widths(ws, [28, 58, 48, 13, 11])
    ws.freeze_panes = "B2"


def build_fusions(ws):
    headers = ["启用", "文件", "配方ID", "原料卡1 ID", "原料卡1名称", "原料卡2 ID", "原料卡2名称", "结果方式",
               "结果卡文件", "结果卡ID", "结果名称", "结果费用", "费用说明"]
    ws.append(headers)
    rows = ft.parse_all()
    for row in rows:
        ws.append(["是" if row["enabled"] else "否", row["file"], row["recipe_id"], row["first_id"], row["first_name"], row["second_id"],
                   row["second_name"], row["mode"], row["result_path"], row["result_id"], row["result_name"],
                   row["result_cost"], row["cost_note"]])
    # 预留空白行，用户直接在表尾填写配方 ID、两张原料卡 ID 和结果设置即可新增。
    for _ in range(20):
        ws.append(["是", "", "", "", "", "", "", "动态合成", "", "", "", -99,
                   "新增行：填写配方ID与两张原料卡ID；-99 表示自动费用"])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), {2, 5, 7, 10, 13})
    for index, row in enumerate(rows, 2):
        if row["mode"] == "固定成品卡":
            ws.cell(index, 11).fill = READONLY_FILL
            ws.cell(index, 12).fill = READONLY_FILL
    _dv(ws, ["是", "否"], 1, ws.max_row)
    _dv(ws, ["动态合成", "固定成品卡"], 8, ws.max_row)
    card_last_row = len(ct.find_card_files()) + 1
    for col in (4, 6):
        _dv_formula(ws, f"'卡牌'!$B$2:$B${card_last_row}", col, ws.max_row)
    _widths(ws, [8, 49, 27, 23, 15, 23, 15, 13, 43, 24, 18, 10, 46])
    ws.freeze_panes = "D2"


def build_blessings(ws):
    headers = ["来源", "来源描述", "祝福名", "祝福描述", "命格", "图标"]
    for i in range(1, bt.MAX_EFFECTS + 1):
        headers += [f"效果{i}", f"效果{i}值"]
    ws.append(headers)
    readonly = {1, 2}
    for r in bt.to_rows():
        row = [r["source"], r["source_desc"], r["name"], r["desc"],
               labels.CLASS_LABELS.get(r["cls"], r["cls"]), labels.ICON_LABELS.get(r["icon"], r["icon"])]
        for i in range(bt.MAX_EFFECTS):
            if i < len(r["effects"]):
                t, a = r["effects"][i]
                row += [labels.BLESSING_EFFECT_LABELS.get(t, t), a]
            else:
                row += ["", None]
        ws.append(row)
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), readonly)
    last = ws.max_row
    _dv(ws, list(labels.CLASS_LABELS.values()), 5, last)
    _dv(ws, list(labels.ICON_LABELS.values()), 6, last)
    for i in range(bt.MAX_EFFECTS):
        _dv(ws, [""] + list(labels.BLESSING_EFFECT_LABELS.values()), 7 + i * 2, last)
    _widths(ws, [12, 30, 14, 42, 10, 12] + [14, 8] * bt.MAX_EFFECTS)
    ws.freeze_panes = "C2"


def build_ai(ws):
    headers = ["行动配置文件", "使用怪物", "行动ID", "行动名称", "行动脚本文件", "选择方式", "出招序列"]
    for i in range(1, at.MAX_PARAMS + 1):
        headers += [f"参数{i}", f"参数{i}值"]
    ws.append(headers)
    readonly = {1, 2, 3, 4, 5, 6}
    readonly.update(8 + i * 2 for i in range(at.MAX_PARAMS))
    for ai in at.parse_all():
        for i, a in enumerate(ai["actions"]):
            row = [ai["file"], ai["users"], a["node"], labels.AI_ACTION_LABELS.get(a["script"], "未命名行动"),
                   a["script_path"], a["type"],
                   ai["sequence"] if i == 0 else ""]
            for j in range(at.MAX_PARAMS):
                param = a["params"][j] if j < len(a["params"]) else {"name": "", "value": None}
                row += [labels.PARAM_LABELS.get(param["name"], param["name"]), param["value"]]
            ws.append(row)
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), readonly)
    _widths(ws, [52, 30, 18, 20, 52, 9, 18] + [21, 10] * at.MAX_PARAMS)
    ws.freeze_panes = "C2"


def build_enemies(ws):
    headers = ["文件", "ID", "名称", "介绍", "最大生命", "行动配置文件", "初始状态ID", "二阶段阈值",
               "二阶段名称", "二阶段护体", "二阶段伤害加成", "二阶段序列"]
    ws.append(headers)
    for e in ent.parse_all():
        ws.append([e["file"], e["id"], e["name"], e["description"], e["max_health"], e["ai_file"],
                   e["starting_statuses"], e["phase_ratio"], e["phase_name"], e["phase_block"],
                   e["phase_damage_bonus"], e["phase_sequence"]])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), {1, 2, 6, 7})
    for col in (8, 11):
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].number_format = "0%"
    _widths(ws, [58, 24, 16, 42, 10, 54, 28, 12, 18, 12, 16, 20])
    ws.freeze_panes = "C2"


def build_levels(ws):
    headers = ["文件", "关卡ID", "类型", "章节起", "章节止", "权重", "金币下限", "金币上限",
               "生命倍率", "伤害倍率", "敌人数", "敌人1", "敌人2", "敌人3", "编队预览", "场景"]
    ws.append(headers)
    for level in lt.parse_all():
        ids = [e["id"] for e in level["enemies"]]
        ws.append([level["file"], level["id"], level["tier"], level["chapter_min"], level["chapter_max"],
                   level["weight"], level["gold_min"], level["gold_max"], level["health_multiplier"],
                   level["damage_multiplier"], level["enemy_count"],
                   *[ids[i] if i < len(ids) else "" for i in range(lt.MAX_ENEMIES)],
                   level["encounter"], level["scene"]])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), {1, 2, 11, 15, 16})
    _dv(ws, list(lt.TIER.values()), 3, ws.max_row)
    for col in (4, 5):
        _dv(ws, ["1", "2", "3"], col, ws.max_row)
    for col in (12, 13, 14):
        _dv_formula(ws, "'怪物'!$B$2:$B$%s" % (len(ent.enemy_by_id()) + 1), col, ws.max_row)
    _widths(ws, [50, 30, 9, 8, 8, 9, 11, 11, 11, 11, 9, 24, 24, 24, 40, 50])
    ws.freeze_panes = "C2"


def build_events(ws):
    headers = ["章节", "文件", "标题", "正文", "插画"]
    for i in range(1, et.MAX_CHOICES + 1):
        headers += [f"选项{i}文案", f"选项{i}效果", f"选项{i}数值"]
    ws.append(headers)
    for event in et.parse_all():
        row = [event["chapter"], event["file"], event["title"], event["body"], event["illustration"]]
        for choice in event["choices"]:
            row += [choice["text"], labels.event_effect_to_zh(choice["effect"]), choice["amount"]]
        ws.append(row)
    readonly = {1, 2, 5}
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), readonly)
    _widths(ws, [7, 50, 18, 46, 42] + [34, 24, 8] * et.MAX_CHOICES)
    ws.freeze_panes = "C2"


def build_event_roots(ws):
    headers = ["章节", "文件", "事件", "覆盖选项"]
    for name in et.ROOT_NAMES[1:]:
        headers += [f"{name}文案", f"{name}效果", f"{name}数值"]
    ws.append(headers)
    for event in et.parse_all():
        if event["root_choice_index"] < 0:
            continue
        row = [event["chapter"], event["file"], event["title"], event["root_choice_index"] + 1]
        for i in range(1, 6):
            row += [event["root_texts"][i] if i < len(event["root_texts"]) else "",
                    labels.event_effect_to_zh(event["root_effects"][i]) if i < len(event["root_effects"]) else "",
                    event["root_amounts"][i] if i < len(event["root_amounts"]) else 0]
        ws.append(row)
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), {1, 2, 3})
    _dv(ws, ["1", "2", "3"], 4, ws.max_row)
    _widths(ws, [7, 50, 18, 10] + [34, 24, 8] * 5)
    ws.freeze_panes = "D2"


def build_potions(ws):
    headers = ["文件", "ID", "名称", "类别", "稀有度", "目标", "战斗外可用", "职业", "说明"]
    for i in range(1, 3):
        headers += [f"效果{i}", f"效果{i}参数1", f"效果{i}值1", f"效果{i}参数2", f"效果{i}值2"]
    ws.append(headers)
    for potion in pt.parse_all():
        row = [potion["file"], potion["id"], potion["name"], potion["category"], potion["rarity"],
               potion["target"], potion["out_of_combat"], potion["character"], potion["tooltip"]]
        for i in range(2):
            effect = potion["effects"][i] if i < len(potion["effects"]) else {"label": "", "params": []}
            params = effect["params"]
            row += [effect["label"], labels.PARAM_LABELS.get(params[0]["name"], params[0]["name"]) if params else "", params[0]["value"] if params else None,
                    labels.PARAM_LABELS.get(params[1]["name"], params[1]["name"]) if len(params) > 1 else "", params[1]["value"] if len(params) > 1 else None]
        ws.append(row)
    readonly = {1, 2, 10, 11, 13, 15, 16, 18}
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), readonly)
    _dv(ws, list(pt.CATEGORY.values()), 4, ws.max_row)
    _dv(ws, list(pt.RARITY.values()), 5, ws.max_row)
    _dv(ws, list(pt.TARGET.values()), 6, ws.max_row)
    _dv(ws, ["是", "否"], 7, ws.max_row)
    _dv(ws, list(pt.CHARACTER.values()), 8, ws.max_row)
    _widths(ws, [48, 22, 14, 8, 7, 10, 11, 9, 44] + [14, 13, 8, 13, 8] * 2)
    ws.freeze_panes = "C2"


def build_shop_prices(ws):
    headers = ["类别", "档位", "基价", "最低价", "最高价", "比例", "定位键", "说明"]
    ws.append(headers)
    for row in st.parse_all():
        ws.append([row["group"], row["tier"], row["base"], row["min"], row["max"], row["percent"], row["key"], row["note"]])
    _style_header(ws, len(headers))
    _style_body(ws, len(headers), {1, 2, 7, 8})
    for cell in ws["F"][1:]:
        cell.number_format = "0%"
    _widths(ws, [12, 12, 10, 10, 10, 10, 22, 38])
    ws.freeze_panes = "C2"


def build_legend(ws):
    notes = [
        "万劫求仙 数值总表 — 使用说明",
        "",
        "1. 只改【浅黄色】列；【灰色】列是只读定位/分组用，尽量别动。",
        "2. 改完保存后运行回写脚本：scripts/xlsx_to_game_data.py",
        "3. 然后让游戏编辑器导入一次资源（打开项目即可）。",
        "",
        "【卡牌】可调基础信息与全部关键词开关；检索/取回/归墟还需填写 1~10 张。",
        "【卡牌效果】每行只对应一个可复用效果组件的一个数值参数，可分别调整自损、伤害、抽牌等效果。",
        "【关键词标准】列出关键词语义与冲突优先级；【效果标准】汇总实际效果组件。",
        "【卡牌】突破方向可选择不可突破、数值提高50%或费用减少1。",
        "【融合配方】现有配方可改原料、结果方式、名称和费用；在表尾空白行填写配方ID与两张原料卡ID即可新增。",
        "   启用填‘否’会从游戏融合库停用但保留配方文件；动态费用填 -99 表示按两张牌费用自动计算。",
        "【祝福】可改文案、命格、图标、效果类型和数值；导入时按表整体重建数据文件。",
        "【事件】可改标题、正文、三个选项文案、效果和数值；组合效果写成 失去生命:8|获得灵石:50。",
        "【事件灵根】只列出 3 个带五行专属分支的事件；覆盖选项填 1/2/3。",
        "【符箓丹药】可改基础信息、说明和现有效果参数值；效果类型/参数名为只读结构。",
        "【怪物】覆盖全部 39 个怪物，可调名称、介绍、生命与二阶段参数；行动配置文件和初始状态只读。",
        "【怪物行动】覆盖全部 27 套配置；使用怪物列会标明共享者，修改共享配置会影响所有列出的怪物。",
        "   出招序列: 子节点下标(从0)按回合循环，如 0,1,2 = 第1回合行动0 → 第2回合行动1 → ...",
        "   格挡行动改数值时，意图显示文字会自动同步。",
        "【关卡设计】覆盖当前战斗池 51 个关卡，可调章节、权重、奖励、倍率，并在现有槽位内替换怪物。",
        "   敌人数为只读；当前不支持直接增减编队槽位。",
        "【商店价格】基价/区间/比例按行填写；比例用 10% 或 0.10 均可。",
    ]
    for n in notes:
        ws.append([n])
    ws.column_dimensions["A"].width = 96
    ws["A1"].font = Font(name=FONT, bold=True, size=13)


def main() -> None:
    wb = Workbook()
    ws_cards = wb.active
    ws_cards.title = "卡牌"
    build_cards(ws_cards)
    build_card_effects(wb.create_sheet("卡牌效果"))
    build_keyword_standards(wb.create_sheet("关键词标准"))
    build_effect_standards(wb.create_sheet("效果标准"))
    build_fusions(wb.create_sheet("融合配方"))
    build_blessings(wb.create_sheet("祝福"))
    build_events(wb.create_sheet("事件"))
    build_event_roots(wb.create_sheet("事件灵根"))
    build_potions(wb.create_sheet("符箓丹药"))
    build_enemies(wb.create_sheet("怪物"))
    build_ai(wb.create_sheet("怪物行动"))
    build_levels(wb.create_sheet("关卡设计"))
    build_shop_prices(wb.create_sheet("商店价格"))
    build_legend(wb.create_sheet("说明"))
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
