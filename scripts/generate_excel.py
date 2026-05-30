import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cards"

headers = [
    "ID", "Display Name", "Character", "Card Type", "Rarity", "Target",
    "Cost", "Exhausts", "Base Damage", "Base Block", "Cards To Draw",
    "Muscle Stacks", "Qi Flow Stacks", "Exposed Duration", "Self Damage",
    "Extra Damage", "Effect Text", "Script Class", "Icon", "Sound"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2f5496", end_color="2f5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
cell_alignment = Alignment(vertical="center", wrap_text=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

cards = [
    # Beastmaster
    ["beast_bond", "万兽同心", "beastmaster", "POWER", "RARE", "SELF",
     1, False, 0, 3, 0, 0, 1, 0, 0, 0,
     "获得 {block} 点护体和 {qi} 层灵息。", "CultivationCard", "tile_0122.png", "true_strength.ogg"],
    ["spirit_crane", "灵鹤引", "beastmaster", "SKILL", "COMMON", "SELF",
     1, False, 0, 0, 2, 0, 0, 0, 0, 0,
     "抽 {draw} 张牌。", "CultivationCard", "draw.png", ""],
    ["tiger_claw", "虎爪", "beastmaster", "ATTACK", "COMMON", "SINGLE_ENEMY",
     1, False, 6, 0, 0, 0, 0, 0, 0, 0,
     "造成 {damage} 点伤害。", "CultivationCard", "tile_0120.png", "slash.ogg"],
    ["turtle_shell", "玄龟甲", "beastmaster", "SKILL", "COMMON", "SELF",
     1, False, 0, 7, 0, 0, 0, 0, 0, 0,
     "获得 {block} 点护体。", "CultivationCard", "tile_0129.png", "block.ogg"],
    ["wolf_pack", "群狼逐月", "beastmaster", "ATTACK", "UNCOMMON", "ALL_ENEMIES",
     2, False, 3, 0, 1, 0, 0, 0, 0, 0,
     "对所有敌人造成 {damage} 点伤害，抽 {draw} 张牌。", "CultivationCard", "tile_0121.png", "slash.ogg"],

    # Body Cultivator
    ["bone_tempering", "淬骨", "body_cultivator", "SKILL", "UNCOMMON", "SELF",
     0, True, 0, 3, 1, 0, 0, 0, 0, 0,
     "获得 {block} 点护体，抽 {draw} 张牌。", "CultivationCard", "tile_0127.png", "block.ogg"],
    ["collapsing_palm", "崩山掌", "body_cultivator", "ATTACK", "UNCOMMON", "SINGLE_ENEMY",
     2, False, 9, 0, 0, 0, 0, 1, 0, 0,
     "造成 {damage} 点伤害，并施加 {exposed} 回合破绽。", "CultivationCard", "tile_0126.png", "axe.ogg"],
    ["iron_bone_fist", "铁骨拳", "body_cultivator", "ATTACK", "COMMON", "SINGLE_ENEMY",
     1, False, 6, 0, 0, 0, 0, 0, 0, 0,
     "造成 {damage} 点伤害。", "CultivationCard", "tile_0118.png", "slash.ogg"],
    ["mountain_breath", "山息", "body_cultivator", "SKILL", "COMMON", "SELF",
     1, False, 0, 6, 0, 0, 0, 0, 0, 0,
     "获得 {block} 点护体。", "CultivationCard", "tile_0129.png", "block.ogg"],
    ["vajra_body", "金刚不坏", "body_cultivator", "POWER", "RARE", "SELF",
     1, False, 0, 0, 0, 0, 2, 0, 0, 0,
     "获得 {qi} 层灵息。灵息会在回合开始转化为劲气。", "CultivationCard", "tile_0127.png", "true_strength.ogg"],

    # Demonic Cultivator
    ["blood_blade", "血刃", "demonic_cultivator", "ATTACK", "COMMON", "SINGLE_ENEMY",
     0, False, 7, 0, 0, 0, 0, 0, 2, 0,
     "失去 {self_damage} 点生命，造成 {damage} 点伤害。", "CultivationCard", "tile_0103.png", "slash.ogg"],
    ["demon_flame", "魔焰焚心", "demonic_cultivator", "ATTACK", "UNCOMMON", "ALL_ENEMIES",
     2, False, 6, 0, 0, 0, 0, 0, 3, 0,
     "失去 {self_damage} 点生命，对所有敌人造成 {damage} 点伤害。", "CultivationCard", "tile_0104.png", "axe.ogg"],
    ["forbidden_mantra", "禁咒入魔", "demonic_cultivator", "POWER", "RARE", "SELF",
     1, False, 0, 0, 0, 0, 2, 0, 3, 0,
     "失去 {self_damage} 点生命，获得 {qi} 层灵息。", "CultivationCard", "tile_0128.png", "true_strength.ogg"],
    ["shadow_step", "影遁", "demonic_cultivator", "SKILL", "UNCOMMON", "SELF",
     0, True, 0, 0, 2, 0, 0, 0, 0, 0,
     "抽 {draw} 张牌。", "CultivationCard", "draw.png", ""],
    ["soul_drain", "摄魂", "demonic_cultivator", "ATTACK", "COMMON", "SINGLE_ENEMY",
     1, False, 5, 3, 0, 0, 0, 0, 0, 0,
     "造成 {damage} 点伤害，获得 {block} 点护体。", "CultivationCard", "tile_0108.png", "slash.ogg"],

    # Sword Cultivator
    ["flying_sword", "飞剑", "sword_cultivator", "ATTACK", "COMMON", "SINGLE_ENEMY",
     1, False, 5, 0, 0, 0, 0, 0, 0, 0,
     "造成 {damage} 点伤害。", "CultivationCard", "tile_0119.png", "slash.ogg"],
    ["guard_sword", "御剑守心", "sword_cultivator", "SKILL", "COMMON", "SELF",
     1, False, 0, 4, 1, 0, 0, 0, 0, 0,
     "获得 {block} 点护体，抽 {draw} 张牌。", "CultivationCard", "tile_0129.png", "block.ogg"],
    ["piercing_light", "一线天", "sword_cultivator", "ATTACK", "UNCOMMON", "SINGLE_ENEMY",
     2, False, 9, 0, 0, 0, 0, 2, 0, 0,
     "造成 {damage} 点伤害，施加 {exposed} 回合破绽。", "CultivationCard", "tile_0130.png", "axe.ogg"],
    ["sword_heart", "剑心通明", "sword_cultivator", "POWER", "RARE", "SELF",
     1, False, 0, 0, 0, 1, 1, 0, 0, 0,
     "获得 {qi} 层灵息和 {muscle} 层劲气。", "CultivationCard", "tile_0131.png", "true_strength.ogg"],
    ["sword_rain", "剑雨", "sword_cultivator", "ATTACK", "COMMON", "ALL_ENEMIES",
     2, False, 4, 0, 0, 0, 0, 0, 0, 0,
     "对所有敌人造成 {damage} 点伤害。", "CultivationCard", "tile_0118.png", "slash.ogg"],

    # Warrior
    ["warrior_angry_anvil", "怒砧", "warrior", "SKILL", "UNCOMMON", "SELF",
     1, False, 0, 7, 0, 0, 0, 0, 0, 0,
     "获得 {block} 点护体，随机消耗 1 张手牌。", "WarriorCard", "tile_0074.png", "block.ogg"],
    ["warrior_axe_attack", "斧劈", "warrior", "ATTACK", "COMMON", "SINGLE_ENEMY",
     1, False, 6, 0, 0, 0, 0, 0, 0, 0,
     "造成 {damage} 点伤害。", "WarriorCard", "tile_0119.png", "axe.ogg"],
    ["warrior_big_slam", "重击", "warrior", "ATTACK", "UNCOMMON", "SINGLE_ENEMY",
     2, True, 4, 0, 0, 0, 0, 2, 0, 0,
     "造成 {damage} 点伤害，并施加 {exposed} 回合破绽。", "WarriorCard", "tile_0117.png", "slash.ogg"],
    ["warrior_block", "格挡", "warrior", "SKILL", "COMMON", "SELF",
     1, False, 0, 5, 0, 0, 0, 0, 0, 0,
     "获得 {block} 点护体。", "WarriorCard", "tile_0102.png", "block.ogg"],
    ["warrior_sharp_knife", "磨刃", "warrior", "ATTACK", "RARE", "SINGLE_ENEMY",
     1, False, 6, 0, 0, 0, 0, 0, 0, 4,
     "造成 {damage} 点伤害。本场战斗此牌伤害 +{extra}。", "WarriorCard", "tile_0105.png", "slash.ogg"],
    ["warrior_slash", "横扫", "warrior", "ATTACK", "COMMON", "ALL_ENEMIES",
     2, False, 4, 0, 0, 0, 0, 0, 0, 0,
     "对所有敌人造成 {damage} 点伤害。", "WarriorCard", "tile_0118.png", "slash.ogg"],
    ["warrior_trickster", "佯攻", "warrior", "SKILL", "UNCOMMON", "SELF",
     0, False, 0, 2, 1, 0, 0, 0, 0, 0,
     "获得 {block} 点护体，抽 1 张牌。", "WarriorCard", "tile_0101.png", "enemy_block.ogg"],
    ["warrior_true_strength", "真武形", "warrior", "POWER", "RARE", "SELF",
     3, False, 0, 0, 0, 0, 0, 0, 0, 0,
     "回合开始时，获得 2 层劲气。", "WarriorCard", "tile_0127.png", "true_strength.ogg"],

    # Common
    ["toxin", "浊气", "common", "SKILL", "COMMON", "SELF",
     1, True, 0, 0, 0, 0, 0, 0, 0, 0,
     "拖慢你的行动。", "Card", "tile_0114.png", "true_strength.ogg"],
]

for row_idx, card in enumerate(cards, 2):
    for col_idx, val in enumerate(card, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = cell_alignment

type_fills = {
    "ATTACK": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "SKILL": PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid"),
    "POWER": PatternFill(start_color="FFF8E0", end_color="FFF8E0", fill_type="solid"),
}

for row_idx in range(2, len(cards) + 2):
    card_type = ws.cell(row=row_idx, column=4).value
    if card_type in type_fills:
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = type_fills[card_type]

col_widths = [22, 14, 20, 12, 12, 16, 8, 10, 14, 14, 16, 14, 14, 16, 14, 14, 55, 18, 22, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

dv_type = DataValidation(type="list", formula1='"ATTACK,SKILL,POWER"', allow_blank=True)
dv_type.error = "Please select a valid card type"
dv_type.errorTitle = "Invalid Card Type"
ws.add_data_validation(dv_type)
dv_type.add(f"D2:D{len(cards)+1}")

dv_rarity = DataValidation(type="list", formula1='"COMMON,UNCOMMON,RARE"', allow_blank=True)
dv_rarity.error = "Please select a valid rarity"
dv_rarity.errorTitle = "Invalid Rarity"
ws.add_data_validation(dv_rarity)
dv_rarity.add(f"E2:E{len(cards)+1}")

dv_target = DataValidation(type="list", formula1='"SELF,SINGLE_ENEMY,ALL_ENEMIES,EVERYONE"', allow_blank=True)
dv_target.error = "Please select a valid target"
dv_target.errorTitle = "Invalid Target"
ws.add_data_validation(dv_target)
dv_target.add(f"F2:F{len(cards)+1}")

chars = "beastmaster,body_cultivator,demonic_cultivator,sword_cultivator,warrior,common"
dv_char = DataValidation(type="list", formula1=f'"{chars}"', allow_blank=True)
dv_char.error = "Please select a valid character"
dv_char.errorTitle = "Invalid Character"
ws.add_data_validation(dv_char)
dv_char.add(f"C2:C{len(cards)+1}")

# Notes sheet
ws2 = wb.create_sheet("Field Notes")
notes = [
    ["Field", "Description", "Notes"],
    ["ID", "Unique card identifier", "Must match card .gd/.tres file name"],
    ["Display Name", "Name shown in game UI", ""],
    ["Character", "Which character owns this card", "common = available to all"],
    ["Card Type", "ATTACK/SKILL/POWER", "Determines card behavior"],
    ["Rarity", "COMMON/UNCOMMON/RARE", "Affects draft availability"],
    ["Target", "SELF/SINGLE_ENEMY/ALL_ENEMIES/EVERYONE", "Who card targets"],
    ["Cost", "Mana cost to play", "0 = free"],
    ["Exhausts", "Card removed after use", "TRUE/FALSE"],
    ["Base Damage", "Base damage dealt", "0 if no damage"],
    ["Base Block", "Base shield gained", "0 if no block"],
    ["Cards To Draw", "Cards to draw extra", "0 if no draw"],
    ["Muscle Stacks", "Muscle status stacks", "Muscle increases damage"],
    ["Qi Flow Stacks", "Qi Flow status stacks", "Qi converts to Muscle at turn start"],
    ["Exposed Duration", "Exposed debuff turns", "Exposed = take more damage"],
    ["Self Damage", "HP lost when used", "0 = no self damage"],
    ["Extra Damage", "Extra damage (磨刃 only)", "Only for warrior_sharp_knife"],
    ["Effect Text", "Description with {placeholders}", "{damage} {block} {draw} {muscle} {qi} {exposed} {self_damage} {extra}"],
    ["Script Class", "Godot script class", "CultivationCard / WarriorCard / Card"],
    ["Icon", "Icon texture file", "Relative to res://art/"],
    ["Sound", "Sound effect file", "Relative to res://art/"],
]

for row_idx, note in enumerate(notes, 1):
    for col_idx, val in enumerate(note, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        cell.border = thin_border

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 45
ws2.column_dimensions["C"].width = 50
ws2.freeze_panes = "A2"

output_path = r"E:\code\game-demo\game-demo\cards_config.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print(f"Total cards: {len(cards)}")
