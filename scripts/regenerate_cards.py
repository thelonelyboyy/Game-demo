#!/usr/bin/env python3
"""
Regenerate Godot .tres card files from cards_config.xlsx.
Usage: python regenerate_cards.py [--dry-run]
"""

import openpyxl
import os
import sys
from pathlib import Path

BASE_DIR = Path(r"E:\code\game-demo\game-demo")
EXCEL_PATH = BASE_DIR / "cards_config.xlsx"

TYPE_MAP = {"ATTACK": 0, "SKILL": 1, "POWER": 2}
RARITY_MAP = {"COMMON": 0, "UNCOMMON": 1, "RARE": 2}
TARGET_MAP = {"SELF": 0, "SINGLE_ENEMY": 1, "ALL_ENEMIES": 2, "EVERYONE": 3}

SCRIPT_CLASS_MAP = {
    "CultivationCard": "res://custom_resources/cultivation_card.gd",
    "Card": "res://custom_resources/card.gd",
}

def get_tres_output_path(card_id, character):
    if character == "common":
        return BASE_DIR / "common_cards" / f"{card_id}.tres"
    else:
        return BASE_DIR / "characters" / character / "cards" / f"{card_id}.tres"


def generate_cultivation_tres(card_data):
    lines = []
    load_steps = 2
    ext_resources = [('type="Script"', f'path="{SCRIPT_CLASS_MAP["CultivationCard"]}"', 'id="1"')]
    next_ext_id = 2

    if card_data.get("icon"):
        ext_resources.append(('type="Texture2D"', f'path="res://art/{card_data["icon"]}"', f'id="{next_ext_id}"'))
        next_ext_id += 1
        load_steps += 1

    if card_data.get("sound"):
        ext_resources.append(('type="AudioStream"', f'path="res://art/{card_data["sound"]}"', f'id="{next_ext_id}"'))
        next_ext_id += 1
        load_steps += 1

    lines.append(f'[gd_resource type="Resource" script_class="CultivationCard" load_steps={load_steps} format=3]')
    lines.append("")

    for res in ext_resources:
        lines.append(f'[ext_resource {" ".join(res)}]')

    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1")')

    cultivation_fields = [
        ("base_damage", "base_damage"),
        ("base_block", "base_block"),
        ("cards_to_draw", "cards_to_draw"),
        ("muscle_stacks", "muscle_stacks"),
        ("qi_flow_stacks", "qi_flow_stacks"),
        ("exposed_duration", "exposed_duration"),
        ("self_damage", "self_damage"),
    ]

    for excel_key, tres_key in cultivation_fields:
        val = card_data.get(excel_key, 0)
        if val != 0:
            lines.append(f"{tres_key} = {val}")

    if card_data.get("effect_text"):
        escaped = card_data["effect_text"].replace('"', '\\"')
        lines.append(f'effect_text = "{escaped}"')

    lines.append(f'id = "{card_data["id"]}"')
    lines.append(f"type = {TYPE_MAP[card_data['card_type']]}")
    lines.append(f"rarity = {RARITY_MAP[card_data['rarity']]}")
    lines.append(f"target = {TARGET_MAP[card_data['target']]}")
    lines.append(f'cost = {card_data["cost"]}')

    if card_data.get("exhausts"):
        lines.append("exhausts = true")

    lines.append(f'display_name = "{card_data["display_name"]}"')

    if card_data.get("icon"):
        lines.append('icon = ExtResource("2")')

    if card_data.get("sound"):
        sound_id = 2 if not card_data.get("icon") else 3
        lines.append(f'sound = ExtResource("{sound_id}")')

    return "\n".join(lines) + "\n"


def generate_base_card_tres(card_data):
    load_steps = 2
    ext_resources = []
    next_ext_id = 1

    ext_resources.append(('type="Script"', f'path="{SCRIPT_CLASS_MAP["Card"]}"', f'id="{next_ext_id}"'))
    next_ext_id += 1
    load_steps += 1

    if card_data.get("icon"):
        ext_resources.append(('type="Texture2D"', f'path="res://art/{card_data["icon"]}"', f'id="{next_ext_id}"'))
        next_ext_id += 1
        load_steps += 1

    if card_data.get("sound"):
        ext_resources.append(('type="AudioStream"', f'path="res://art/{card_data["sound"]}"', f'id="{next_ext_id}"'))
        next_ext_id += 1
        load_steps += 1

    lines = [f'[gd_resource type="Resource" script_class="Card" load_steps={load_steps} format=3]']
    lines.append("")

    for res in ext_resources:
        lines.append(f'[ext_resource {" ".join(res)}]')

    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1")')

    lines.append(f'id = "{card_data["id"]}"')
    lines.append(f"type = {TYPE_MAP[card_data['card_type']]}")
    lines.append(f"rarity = {RARITY_MAP[card_data['rarity']]}")
    lines.append(f"target = {TARGET_MAP[card_data['target']]}")
    lines.append(f'cost = {card_data["cost"]}')

    if card_data.get("exhausts"):
        lines.append("exhausts = true")
    else:
        lines.append("exhausts = false")

    if card_data.get("icon"):
        lines.append('icon = ExtResource("2")')

    lines.append(f'display_name = "{card_data["display_name"]}"')

    tt = card_data.get("effect_text", "")
    escaped = tt.replace('"', '\\"')
    lines.append(f'tooltip_text = "[center]{escaped}[/center]"')

    if card_data.get("sound"):
        sound_id = 3 if card_data.get("icon") else 2
        lines.append(f'sound = ExtResource("{sound_id}")')

    return "\n".join(lines) + "\n"


def read_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Cards"]

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    cards = []
    for row in range(2, ws.max_row + 1):
        card = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                val = "" if header in ("Effect Text", "Display Name", "Icon", "Sound", "Script Class") else 0
            elif header == "Exhausts":
                val = str(val).upper() == "TRUE"
            elif header in ("Cost", "Base Damage", "Base Block", "Cards To Draw",
                             "Muscle Stacks", "Qi Flow Stacks", "Exposed Duration",
                             "Self Damage", "Extra Damage"):
                val = int(val) if val else 0
            card[header.lower().replace(" ", "_")] = val

        card["id"] = card["id"]
        cards.append(card)

    wb.close()
    return cards


def main():
    dry_run = "--dry-run" in sys.argv

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        print("Run generate_excel.py first.")
        sys.exit(1)

    print(f"Reading cards from: {EXCEL_PATH}")
    cards = read_excel(EXCEL_PATH)
    print(f"Found {len(cards)} cards.\n")

    generated = 0
    for card in cards:
        card_id = card["id"]
        script_class = card["script_class"]
        character = card["character"]

        out_path = get_tres_output_path(card_id, character)

        if script_class == "CultivationCard":
            content = generate_cultivation_tres(card)
        else:
            content = generate_base_card_tres(card)

        if dry_run:
            print(f"  [DRY-RUN] Would write: {out_path}")
        else:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"  Generated: {out_path}")

        generated += 1

    print(f"\n{'[DRY-RUN] Would have generated' if dry_run else 'Generated'} {generated} .tres files.")

    if not dry_run:
        print("\nDone! In Godot, reload the project to pick up the updated .tres files.")


if __name__ == "__main__":
    main()
