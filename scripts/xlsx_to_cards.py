"""Write edited cards.xlsx values back into the card .tres files.

Run:  python scripts/xlsx_to_cards.py
Only numeric / enum fields are written: 名称/费用/类型/目标/稀有度/消耗/元素 and
each effect's 值 (amount). Effect types and structure are left untouched.
After running, reimport in the Godot editor.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import card_table as ct

XLSX = Path(__file__).resolve().parents[1] / "cards.xlsx"


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"not found: {XLSX} (run cards_to_xlsx.py first)")

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["cards"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    changed = 0
    errors = []
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if not vals or not vals[idx["文件"]]:
            continue
        rel = str(vals[idx["文件"]]).strip()
        path = ct.PROJECT_ROOT / rel
        if not path.exists():
            errors.append(f"row {r}: file not found {rel}")
            continue

        def gv(name, default=""):
            v = vals[idx[name]]
            return default if v is None else v

        try:
            row = {
                "name": str(gv("名称")),
                "cost": int(gv("费用", 0)),
                "type": str(gv("类型")).strip(),
                "target": str(gv("目标")).strip(),
                "rarity": str(gv("稀有度")).strip(),
                "exhausts": str(gv("消耗")).strip() == "是",
                "element": str(gv("元素")).strip(),
                "effect_amounts": [],
            }
            # validate enum membership
            for key, table in (("type", ct.TYPE_R), ("target", ct.TARGET_R),
                               ("rarity", ct.RARITY_R), ("element", ct.ELEMENT_R)):
                if row[key] not in table:
                    raise ValueError(f"非法{key}值 '{row[key]}'")
            for i in range(1, ct.MAX_EFFECTS + 1):
                label = vals[idx[f"效果{i}"]]
                amt = vals[idx[f"效果{i}值"]]
                if label in (None, ""):
                    row["effect_amounts"].append(None)
                else:
                    row["effect_amounts"].append(int(amt) if amt is not None else None)
        except (ValueError, TypeError) as e:
            errors.append(f"row {r} ({rel}): {e}")
            continue

        if ct.update_card_file(path, row):
            changed += 1

    print(f"updated {changed} card file(s)")
    if errors:
        print("跳过(有错误):")
        for e in errors:
            print("  " + e)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
